from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from tkinter import BOTH, BOTTOM, END, LEFT, RIGHT, TOP, X, YES, Canvas, Frame, Label, Entry, Button, Tk, StringVar, Text, Scrollbar, messagebox
from tkinter import ttk

from openpyxl import Workbook, load_workbook


APP_TITLE = "E-Tricycle Portal"
WINDOW_SIZE = "1040x720"

BG = "#0f172a"
PANEL = "#111827"
SURFACE = "#1f2937"
SURFACE_2 = "#273244"
TEXT = "#e5e7eb"
MUTED = "#94a3b8"
ACCENT = "#38bdf8"
ACCENT_2 = "#22c55e"
WARNING = "#f59e0b"
DANGER = "#ef4444"
USER_BUBBLE = "#2563eb"
BOT_BUBBLE = "#334155"

HEADERS = [
    "Name",
    "Contact details",
    "Vehicle registration No",
    "Chassis No",
    "Email id",
    "Security key",
]


@dataclass
class RegistrationDraft:
    name: str = ""
    contact: str = ""
    vehicle: str = ""
    chassis: str = ""
    email: str = ""
    security: str = ""


class ETricycleApp:
    def __init__(self, root: Tk) -> None:
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry(WINDOW_SIZE)
        self.root.minsize(960, 680)
        self.root.configure(bg=BG)
        self.root.protocol("WM_DELETE_WINDOW", self.on_exit)

        self.base_dir = Path(__file__).resolve().parent
        self.workbook_path = self.base_dir / "E-tricycle.xlsx"
        self.workbook, self.sheet = self._load_or_create_workbook()

        self.draft = RegistrationDraft()
        self.flow: str | None = None
        self.flow_step = 0
        self.status_var = StringVar(value="Ready")
        self.input_var = StringVar()

        self._build_theme()
        self._build_ui()
        self._show_welcome()

    def _build_theme(self) -> None:
        style = ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure("Header.TLabel", background=BG, foreground=TEXT, font=("Segoe UI", 24, "bold"))
        style.configure("SubHeader.TLabel", background=BG, foreground=MUTED, font=("Segoe UI", 10))
        style.configure("Panel.TFrame", background=PANEL)
        style.configure("Surface.TFrame", background=SURFACE)
        style.configure("Accent.TLabel", background=SURFACE, foreground=TEXT, font=("Segoe UI", 11, "bold"))

    def _build_ui(self) -> None:
        header = Frame(self.root, bg=BG, padx=26, pady=22)
        header.pack(fill=X)

        title_row = Frame(header, bg=BG)
        title_row.pack(fill=X)
        Label(
            title_row,
            text="E-Tricycle Portal",
            bg=BG,
            fg=TEXT,
            font=("Segoe UI", 24, "bold"),
        ).pack(anchor="w")
        Label(
            title_row,
            text="Chatbot-style registration and login flow with Excel-backed storage",
            bg=BG,
            fg=MUTED,
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(4, 0))

        action_row = Frame(header, bg=BG)
        action_row.pack(fill=X, pady=(18, 0))
        self._add_action_button(action_row, "Register", self.start_registration, ACCENT_2).pack(side=LEFT, padx=(0, 10))
        self._add_action_button(action_row, "Login", self.start_login, ACCENT).pack(side=LEFT, padx=(0, 10))
        self._add_action_button(action_row, "Helpline", self.show_helpline, WARNING).pack(side=LEFT, padx=(0, 10))
        self._add_action_button(action_row, "Reset Chat", self.reset_chat, SURFACE_2).pack(side=LEFT, padx=(0, 10))
        self._add_action_button(action_row, "Exit", self.on_exit, DANGER).pack(side=LEFT)

        main = Frame(self.root, bg=BG, padx=22, pady=0)
        main.pack(fill=BOTH, expand=YES)

        left_panel = Frame(main, bg=PANEL, width=310)
        left_panel.pack(side=LEFT, fill=BOTH, padx=(0, 14), pady=(0, 16))
        left_panel.pack_propagate(False)

        Label(
            left_panel,
            text="Quick Guide",
            bg=PANEL,
            fg=TEXT,
            font=("Segoe UI", 16, "bold"),
        ).pack(anchor="w", padx=18, pady=(18, 8))

        guide_text = (
            "This version behaves like a chatbot.\n\n"
            "1. Click Register and answer step by step.\n"
            "2. Click Login to verify a saved vehicle.\n"
            "3. Use Helpline for contact details.\n"
            "4. Reset Chat if you want a fresh conversation.\n\n"
            "Data is stored in E-tricycle.xlsx next to the script."
        )
        Label(
            left_panel,
            text=guide_text,
            justify=LEFT,
            wraplength=260,
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 10),
            padx=18,
            pady=10,
        ).pack(anchor="w")

        self.metrics_label = Label(
            left_panel,
            text="Excel storage: ready",
            bg=SURFACE,
            fg=TEXT,
            font=("Segoe UI", 10, "bold"),
            padx=12,
            pady=12,
        )
        self.metrics_label.pack(fill=X, padx=18, pady=(14, 18))

        chat_panel = Frame(main, bg=PANEL)
        chat_panel.pack(side=LEFT, fill=BOTH, expand=YES, pady=(0, 16))

        chat_header = Frame(chat_panel, bg=PANEL, padx=18, pady=14)
        chat_header.pack(fill=X)
        Label(
            chat_header,
            text="Conversation",
            bg=PANEL,
            fg=TEXT,
            font=("Segoe UI", 16, "bold"),
        ).pack(anchor="w")
        Label(
            chat_header,
            text="Type a message or use the buttons above.",
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(2, 0))

        chat_body = Frame(chat_panel, bg=PANEL, padx=18, pady=8)
        chat_body.pack(fill=BOTH, expand=YES)

        self.chat_canvas = Canvas(chat_body, bg=PANEL, highlightthickness=0)
        self.chat_scroll = Scrollbar(chat_body, orient="vertical", command=self.chat_canvas.yview)
        self.chat_canvas.configure(yscrollcommand=self.chat_scroll.set)
        self.chat_scroll.pack(side=RIGHT, fill="y")
        self.chat_canvas.pack(side=LEFT, fill=BOTH, expand=YES)

        self.chat_frame = Frame(self.chat_canvas, bg=PANEL)
        self.chat_window = self.chat_canvas.create_window((0, 0), window=self.chat_frame, anchor="nw")
        self.chat_frame.bind("<Configure>", self._on_chat_configure)
        self.chat_canvas.bind("<Configure>", self._on_canvas_configure)

        input_panel = Frame(chat_panel, bg=PANEL, padx=18, pady=16)
        input_panel.pack(fill=X)

        self.entry = Entry(
            input_panel,
            textvariable=self.input_var,
            bg=SURFACE,
            fg=TEXT,
            insertbackground=TEXT,
            relief="flat",
            font=("Segoe UI", 11),
        )
        self.entry.pack(side=LEFT, fill=X, expand=YES, ipady=10, padx=(0, 10))
        self.entry.bind("<Return>", self.send_message)

        self.send_button = Button(
            input_panel,
            text="Send",
            command=self.send_message,
            bg=ACCENT,
            fg="#0f172a",
            activebackground="#7dd3fc",
            activeforeground="#0f172a",
            relief="flat",
            font=("Segoe UI", 10, "bold"),
            padx=18,
            pady=10,
        )
        self.send_button.pack(side=LEFT)

        status_bar = Frame(self.root, bg=SURFACE, height=36)
        status_bar.pack(fill=X, side=BOTTOM)
        Label(
            status_bar,
            textvariable=self.status_var,
            bg=SURFACE,
            fg=TEXT,
            font=("Segoe UI", 10),
            padx=16,
            pady=8,
        ).pack(anchor="w")

    def _add_action_button(self, parent: Frame, text: str, command, color: str) -> Button:
        return Button(
            parent,
            text=text,
            command=command,
            bg=color,
            fg="#0f172a" if color != SURFACE_2 and color != DANGER else TEXT,
            activebackground=color,
            activeforeground="#0f172a" if color != SURFACE_2 and color != DANGER else TEXT,
            relief="flat",
            font=("Segoe UI", 10, "bold"),
            padx=14,
            pady=8,
        )

    def _load_or_create_workbook(self):
        if self.workbook_path.exists():
            workbook = load_workbook(self.workbook_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active

        if sheet.max_row < 1 or [sheet.cell(row=1, column=i).value for i in range(1, 7)] != HEADERS:
            for index, header in enumerate(HEADERS, start=1):
                sheet.cell(row=1, column=index).value = header
            sheet.column_dimensions["A"].width = 30
            sheet.column_dimensions["B"].width = 20
            sheet.column_dimensions["C"].width = 24
            sheet.column_dimensions["D"].width = 24
            sheet.column_dimensions["E"].width = 30
            sheet.column_dimensions["F"].width = 18
            workbook.save(self.workbook_path)

        return workbook, sheet

    def _on_chat_configure(self, _event) -> None:
        self.chat_canvas.configure(scrollregion=self.chat_canvas.bbox("all"))

    def _on_canvas_configure(self, event) -> None:
        self.chat_canvas.itemconfigure(self.chat_window, width=event.width)

    def _scroll_to_bottom(self) -> None:
        self.chat_canvas.update_idletasks()
        self.chat_canvas.yview_moveto(1.0)

    def _show_welcome(self) -> None:
        self._bot_message(
            "Welcome to the E-Tricycle Portal. I can help you register a vehicle, verify a login, or show helpline details."
        )
        self._bot_message("Choose a quick action above, or type Register, Login, or Helpline.")
        self._update_status("Ready to help")
        self._refresh_metrics()
        self.entry.focus_set()

    def _refresh_metrics(self) -> None:
        records = max(0, self.sheet.max_row - 1)
        self.metrics_label.configure(text=f"Excel storage: {records} record(s) saved")

    def _update_status(self, text: str) -> None:
        self.status_var.set(text)

    def _create_bubble(self, sender: str, message: str, bubble_color: str, align_right: bool) -> None:
        outer = Frame(self.chat_frame, bg=PANEL)
        outer.pack(fill=X, pady=6)

        bubble = Frame(outer, bg=bubble_color, padx=14, pady=10)
        bubble.pack(side=RIGHT if align_right else LEFT, anchor="e" if align_right else "w")

        sender_label = Label(
            bubble,
            text=sender,
            bg=bubble_color,
            fg="#dbeafe" if align_right else "#cbd5e1",
            font=("Segoe UI", 9, "bold"),
        )
        sender_label.pack(anchor="w")

        message_label = Label(
            bubble,
            text=message,
            bg=bubble_color,
            fg="#f8fafc",
            justify=LEFT,
            wraplength=560,
            font=("Segoe UI", 10),
        )
        message_label.pack(anchor="w", pady=(4, 0))

        self._scroll_to_bottom()

    def _bot_message(self, message: str) -> None:
        self._create_bubble("Assistant", message, BOT_BUBBLE, align_right=False)

    def _user_message(self, message: str) -> None:
        self._create_bubble("You", message, USER_BUBBLE, align_right=True)

    def _clear_flow(self, keep_chat: bool = False) -> None:
        self.flow = None
        self.flow_step = 0
        self.draft = RegistrationDraft()
        if not keep_chat:
            for child in self.chat_frame.winfo_children():
                child.destroy()
            self._show_welcome()
        else:
            self._update_status("Ready")

    def reset_chat(self) -> None:
        self._clear_flow(keep_chat=False)

    def start_registration(self) -> None:
        self.flow = "register"
        self.flow_step = 0
        self.draft = RegistrationDraft()
        self._update_status("Registration flow active")
        self._bot_message("Let's register a vehicle. What is the owner's name?")
        self.entry.focus_set()

    def start_login(self) -> None:
        self.flow = "login"
        self.flow_step = 0
        self._update_status("Login flow active")
        self._bot_message("Let's verify your account. Please enter the vehicle registration number.")
        self.entry.focus_set()

    def show_helpline(self) -> None:
        self._bot_message("Helpline: +91 1010110 | Email: erc@tricycle.com")
        self._bot_message("If you want, I can also guide you through registration or login right now.")
        self._update_status("Helpline shown")

    def _normalize_vehicle(self, value: str) -> str:
        return re.sub(r"\s+", "", value).upper()

    def _validate_email(self, value: str) -> bool:
        return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", value.strip()))

    def _validate_contact(self, value: str) -> bool:
        cleaned = re.sub(r"[\s\-()]", "", value)
        return bool(re.fullmatch(r"\+?\d{7,15}", cleaned))

    def _find_record_row(self, vehicle: str) -> int | None:
        target = self._normalize_vehicle(vehicle)
        for row in range(2, self.sheet.max_row + 1):
            current = self.sheet.cell(row=row, column=3).value
            if current and self._normalize_vehicle(str(current)) == target:
                return row
        return None

    def _save_registration(self) -> bool:
        if self._find_record_row(self.draft.vehicle) is not None:
            self._bot_message(
                "That vehicle registration number is already saved. Please use a different vehicle number or try Login."
            )
            return False

        next_row = self.sheet.max_row + 1
        values = [
            self.draft.name.strip(),
            self.draft.contact.strip(),
            self._normalize_vehicle(self.draft.vehicle),
            self.draft.chassis.strip(),
            self.draft.email.strip(),
            self.draft.security.strip(),
        ]
        for column, value in enumerate(values, start=1):
            self.sheet.cell(row=next_row, column=column).value = value
        self.workbook.save(self.workbook_path)
        self._refresh_metrics()
        return True

    def _handle_register_step(self, text: str) -> None:
        cleaned = text.strip()
        if self.flow_step == 0:
            if not cleaned:
                self._bot_message("Please enter the owner's name.")
                return
            self.draft.name = cleaned
            self.flow_step = 1
            self._bot_message("Great. Now share the contact number with country code if needed.")
            return

        if self.flow_step == 1:
            if not self._validate_contact(cleaned):
                self._bot_message("That contact number does not look valid. Use digits only, optionally with a leading +.")
                return
            self.draft.contact = cleaned
            self.flow_step = 2
            self._bot_message("Thanks. Enter the vehicle registration number.")
            return

        if self.flow_step == 2:
            if not cleaned:
                self._bot_message("Please enter the vehicle registration number.")
                return
            self.draft.vehicle = self._normalize_vehicle(cleaned)
            self.flow_step = 3
            self._bot_message("Now enter the chassis number.")
            return

        if self.flow_step == 3:
            if not cleaned:
                self._bot_message("Please enter the chassis number.")
                return
            self.draft.chassis = cleaned
            self.flow_step = 4
            self._bot_message("Enter the email address.")
            return

        if self.flow_step == 4:
            if not self._validate_email(cleaned):
                self._bot_message("That email address is not valid. Please try again.")
                return
            self.draft.email = cleaned
            self.flow_step = 5
            self._bot_message("Set a security key for this vehicle.")
            return

        if self.flow_step == 5:
            if not cleaned:
                self._bot_message("Security key cannot be empty.")
                return
            self.draft.security = cleaned
            if self._save_registration():
                self._bot_message(
                    f"Registration successful for {self.draft.name}. Your vehicle {self.draft.vehicle} has been saved."
                )
                self._update_status("Registration completed")
                self._clear_flow(keep_chat=True)
            return

    def _handle_login_step(self, text: str) -> None:
        cleaned = text.strip()
        if self.flow_step == 0:
            if not cleaned:
                self._bot_message("Please enter the vehicle registration number.")
                return
            self.draft.vehicle = self._normalize_vehicle(cleaned)
            record_row = self._find_record_row(self.draft.vehicle)
            if record_row is None:
                self._bot_message(
                    "I could not find that vehicle number in the workbook. Register first if this is a new vehicle."
                )
                self._update_status("Login failed")
                self._clear_flow(keep_chat=True)
                return
            self.flow_step = 1
            self._bot_message("Vehicle found. Now enter the security key.")
            return

        if self.flow_step == 1:
            record_row = self._find_record_row(self.draft.vehicle)
            if record_row is None:
                self._bot_message("The vehicle record is no longer available. Please register again.")
                self._update_status("Login failed")
                self._clear_flow(keep_chat=True)
                return

            saved_key = str(self.sheet.cell(row=record_row, column=6).value or "")
            if cleaned != saved_key:
                self._bot_message("Security key mismatch. Please check the key and try again.")
                self._update_status("Login failed")
                self._clear_flow(keep_chat=True)
                return

            owner = str(self.sheet.cell(row=record_row, column=1).value or "Unknown")
            self._bot_message(f"Login successful. Welcome back, {owner}. Your vehicle profile is verified.")
            self._update_status("Login verified")
            self._clear_flow(keep_chat=True)

    def _handle_command(self, text: str) -> None:
        normalized = text.strip().lower()
        if not normalized:
            return

        commands = {
            "register": self.start_registration,
            "signup": self.start_registration,
            "sign up": self.start_registration,
            "login": self.start_login,
            "sign in": self.start_login,
            "signin": self.start_login,
            "help": self.show_helpline,
            "helpline": self.show_helpline,
            "reset": self.reset_chat,
            "clear": self.reset_chat,
        }

        action = commands.get(normalized)
        if action:
            action()
            return

        self._bot_message(
            "I can help with Register, Login, Helpline, Reset, or Exit. Choose one of those to continue."
        )

    def send_message(self, _event=None) -> None:
        text = self.input_var.get().strip()
        if not text:
            return

        self.input_var.set("")
        self._user_message(text)

        normalized = text.strip().lower()
        if normalized in {
            "register",
            "signup",
            "sign up",
            "login",
            "sign in",
            "signin",
            "help",
            "helpline",
            "reset",
            "clear",
            "exit",
            "quit",
        }:
            self._handle_command(text)
            return

        if self.flow == "register":
            self._handle_register_step(text)
        elif self.flow == "login":
            self._handle_login_step(text)
        else:
            self._handle_command(text)

    def on_exit(self) -> None:
        if messagebox.askyesno(APP_TITLE, "Do you want to close the portal?"):
            self.root.destroy()


def main() -> None:
    root = Tk()
    ETricycleApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
