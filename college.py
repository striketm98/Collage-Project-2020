from tkinter import*
from openpyxl import *
from sys import*
from os import*
from tcl import*
from re import*
from enum import*
path='D:\\College_project2020\\E-tricycle.xlsx'
wb =  load_workbook(path)
sheet = wb.active
current_row=sheet.max_row
current_column=sheet.max_column
def excel():

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 50
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 10    

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Contact details"
    sheet.cell(row=1, column=3).value = "Vehicle registration No" 
    sheet.cell(row=1, column=4).value = "Chais No" 
    sheet.cell(row=1, column=5).value = "Email id" 
    sheet.cell(row=1, column=6).value = "Security key"
    
def help_line_Open():    
    global screen7 
    screen7 = Toplevel(screen) 
    screen7.title("Helpline details") 
    screen7.geometry("400x100") 
    Label(screen7,text="Helpline No: +91 1010110",font=("times new roman",10,"bold")).pack()
    Label(text = "").pack() 
    Label(screen7,text="MAIL US: ERC@TRICYLE.com",font=("times new roman",10,"bold")).pack()  
    Label(text = "").pack() 
    Button(screen7, text="Exit",command=delete6, width = 10, height = 1).pack()

def delete6(): 
    screen7.destroy()
def delete5(): 
    screen2.destroy() 
def delete2(): 
    screen6.destroy() 
def delete3(): 
    screen4.destroy() 
def delete4(): 
    screen5.destroy()

                           
            
def entry_user(): 
    global screen6 
    screen6 = Toplevel(screen) 
    screen6.title("Engine Start") 
    screen6.geometry("250x200") 
    Label(screen6,text="Start Engine",font=("times new roman",15,"bold"),bg= "green").pack()  
    Button(screen6, text="Start").pack() 
    Label(text = "").pack() 
    Label(screen6,text="Stop Engine",font=("times new roman",15,"bold"),bg= "red").pack()  
    Button(screen6, text="Stop",command=delete2).pack() 
  
def login_sucess():
    global screen3
    screen3 = Toplevel(screen)
    screen3.title("Success") 
    screen3.geometry("150x150")
    Label(screen3, text = "Login Sucess").pack()
    Label(screen3,text = "").pack()   
    Button(screen3, text="OK", command =entry_user).pack()

    
def password_not_recognised(): 
    global screen4 
    screen4 = Toplevel(screen) 
    screen4.title("Success") 
    screen4.geometry("150x100") 
    Label(screen4, text = "Security Key Error").pack() 
    Button(screen4, text = "OK", command =delete3).pack()

def user_not_found(): 
    global screen5 
    screen5 = Toplevel(screen)
    screen5.title("Success") 
    screen5.geometry("150x100") 
    Label(screen5, text = "User Not Found").pack() 
    Button(screen5, text = "OK", command =delete4).pack()

def register_user():
    Name_info=Name.get() 
    contact_info=contact.get() 
    vehical_info=vehical.get() 
    chasis_info=chasis.get()  
    email_id_info=email_id.get() 
    security_info=security.get()
    
    
    sheet.cell(row=current_row + 1, column=1).value=Name_info
    sheet.cell(row=current_row + 1, column=2).value=contact_info
    sheet.cell(row=current_row + 1, column=3).value=vehical_info
    sheet.cell(row=current_row + 1, column=4).value=chasis_info
    sheet.cell(row=current_row + 1, column=5).value=email_id_info
    sheet.cell(row=current_row + 1, column=6).value=security_info
    wb.save(path)
    Label(screen1, text = "Registration Sucess", fg = "green" ,font = ("calibri", 11)).pack()


def clear_user(): 
    Name_entry.delete(0, END)  
    contact_entry.delete(0, END) 
    vehical_entry.delete(0, END)
    chasis_entry.delete(0, END) 
    email_id_entry.delete(0, END) 
    security_entry.delete(0, END) 
    Label(screen1, text = "Reset", fg = "red" ,font = ("calibri", 11)).pack() 


def back_user(): 
    Label(screen1, text = "Reset", fg = "red" ,font = ("calibri", 11)).pack()   
    screen1.destroy()


def login_verify(): 
    e_bike1 = e_bike_verify.get() 
    security1=security_verify.get()
    for i in range(1,current_row+1):
        if e_bike1  == sheet.cell(row=i,column=3).value and security1 == sheet.cell(row=i,column=6).value:  
            login_sucess()
        if e_bike1  == sheet.cell(row=i,column=3).value and security1 != sheet.cell(row=i,column=6).value: 
            password_not_recognised()
        if  e_bike1  != sheet.cell(row=i,column=3).value and security1 == sheet.cell(row=i,column=6).value:
            user_not_found()   
     
             
  
def register(): 
    global screen1 
    screen1 = Toplevel(screen)  
    screen1.title("E-tricycle Register")  
    screen1.geometry("350x550")

    global Name 
    global contact  
    global vehical 
    global chasis 
    global email_id 
    global security 
    global Name_entry 
    global contact_entry 
    global vehical_entry 
    global chasis_entry 
    global email_id_entry
    global security_entry

    Name=StringVar() 
    contact=StringVar() 
    vehical=StringVar() 
    chasis=StringVar() 
    email_id=StringVar() 
    security=StringVar() 
  
    Label(screen1, text = "Please enter details below",bg="Grey" ,font=("new times roman",20,"bold")).pack()   
    Label(screen1, text = "").pack() 
  
    Label(screen1, text = "Name").pack() 
    Name_entry = Entry(screen1, textvariable = Name) 
    Name_entry.pack() 
  
  
    Label(screen1, text = "Contact details").pack() 
    contact_entry = Entry(screen1, textvariable = contact) 
    contact_entry.pack()

    Label(screen1, text = "Vehicle registration No").pack() 
    vehical_entry = Entry(screen1, textvariable = vehical) 
    vehical_entry.pack()

    Label(screen1, text = "Chais No").pack() 
    chasis_entry = Entry(screen1, textvariable = chasis) 
    chasis_entry.pack() 
  
    Label(screen1, text = "Email id").pack() 
    email_id_entry = Entry(screen1, textvariable =email_id) 
    email_id_entry.pack() 
  
    Label(screen1, text = "Security key * ").pack() 
    security_entry = Entry(screen1, textvariable =security,show="*") 
    security_entry.pack() 
  
  
    Label(screen1, text = "").pack() 
    Button(screen1, text = "submit", width = 10, height = 1, command = register_user).pack()   
    Label(screen1, text = "").pack() 
    Button(screen1, text = "Reset", width = 10, height = 1, command = clear_user).pack()    
    Label(screen1, text = "").pack()
    Button(screen1, text = "Exit", width = 10, height = 1, command = back_user).pack()


def login(): 
    global screen2 
    screen2 = Toplevel(screen) 
    screen2.title("Login E-tricycle potral") 
    screen2.geometry("350x300") 
    Label(screen2, text = "Please enter your login details", bg="grey",font=("times new roman",20,"bold")).pack()
    Label(screen2,text = "").pack()  
    
    global e_bike_verify 
    global security_verify 
  
    e_bike_verify = StringVar() 
    security_verify = StringVar()

    global e_bike_entry1 
    global security_entry1 
  
    Label(screen2, text = " Vehicle Registration No").pack() 
    e_bike_entry1 = Entry(screen2, textvariable = e_bike_verify) 
    e_bike_entry1.pack() 
     
  
    Label(screen2, text = "Security Key *").pack()
    security_entry1 = Entry(screen2, textvariable = security_verify,show="*") 
    security_entry1.pack()

    Label(screen2,text = "").pack()
    Button(screen2, text = "Login", width = 10, height = 1, command = login_verify).pack()
    Label(screen2,text = "").pack()  
    Button(screen2, text="exit", width = 10, height = 1, command =delete5).pack()
  
  
def main_screen(): 
    global screen 
    screen = Tk() 
    screen.geometry("500x300") 
    screen.title("E-tricycle booking System") 
    Label(text = "E-Tricycle booking System", bg = "grey", width = "300", height = "2", font = ("Calibri", 20,"bold")).pack()  
    Label(text = "").pack() 
    Button(text = "Login", height = "2", width = "30", command = login).pack() 
    Label(text = "").pack() 
    Button(text = "Register",height = "2", width = "30", command = register).pack() 
    Label(text = "").pack() 
    Button(text = "Helpline No",height = "2", width = "30", command = help_line_Open).pack()   
    screen.mainloop() 
main_screen() 
